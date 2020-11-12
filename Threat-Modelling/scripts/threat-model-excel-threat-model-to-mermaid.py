#!/usr/bin/env python3

from openpyxl import load_workbook
import os

# Make sure to set data_only=True, we don't want formulas, just data

workbook = load_workbook(filename="Threat Modelling Spreadsheet v2.0.xlsx",  data_only = "True")

# You can ignore then
#
# /usr/local/lib/python3.8/dist-packages/openpyxl/worksheet/_reader.py:300: UserWarning: Data Validation extension is not supported and will be removed warn(msg)
#
# This occurs because of the data validation used in the workbook which is not supported in openpyxl, and we don't need it since we just want the data

# TODO: iterate through the sheets with relationships like IRelationships, DORelationships, ADRelationships, CBRelationships and ThreatModel

ws_IRelationships = workbook["IRelationships"]

row_count=0

# mermaid data format:
#```mermaid
#graph LR
#BlockchainAccountSecurityModel(BlockchainAccountSecurityModel)
#CredentialStuffingAttack -->|results in|PasswordTheftAttack[Password Theft Attack]

# Initialize some variables here
mermaid_string=""
tags=[]
mermaid_files={}

for row in ws_IRelationships.iter_rows():
    if row_count == 0:
        # Ignore the first row
        # TODO: is there an open csv like function that'll read the sheet in and auto populate a list? For now we use a janky state machine which will break if any columns get moved
        row_count = row_count + 1
    else:
        # Initialize variables at end of each row, the first row isn't used
        cell_count=1
        for cell in row:
            # A, B, C, D, E, F
            # ID,Type,TAGS,Component/InteractionA,Relationship,Component/InteractionB
            if cell_count == 1:
                # skip ID column
                cell_count = cell_count +1
            elif cell_count == 2:
                # skip Type Column
                cell_count = cell_count +1
            elif cell_count == 3:
                # skip TAGS
                tags=cell.value.split(",")
                # Check all tags and create a dict entry if we don't have one
                for tag_entry in tags:
                    # strip the starting space from the tag because, space, entry, for, tags
                    tag_entry=tag_entry.lstrip(' ')
                    if tag_entry not in mermaid_files:
                        mermaid_files[tag_entry]={}
                        mermaid_files[tag_entry]["name"]=tag_entry
                        mermaid_files[tag_entry]["strings"]=[]
                cell_count = cell_count +1
            elif cell_count == 4:
                # Strip all spaces out for named objecty in mermaid
                mermaid_string=cell.value.replace(" ", "") + " -->|"
                cell_count = cell_count +1
            elif cell_count == 5:
                mermaid_string = mermaid_string + cell.value + "|"
                cell_count = cell_count +1
            elif cell_count == 6:
                # Strip all spaces out for named objecty in mermaid but leave spaces for object title
                mermaid_string = mermaid_string + cell.value.replace(" ", "") + "[" + cell.value + "]"
                cell_count = 1
                mermaid_files[tag_entry]["strings"].append(mermaid_string)
                mermaid_string=""
                tags=[]

#
# Put the data into files
#
for key,value in mermaid_files.items():
    outputfilename=key + ".md"
    mermaid_diagram_name=key.replace(" ", "") + "(" + key.replace(" ", "") + ")"
    with open(outputfilename, "w") as outputfile:
        outputfile.write("```mermaid" + os.linesep)
        outputfile.write("LR" + os.linesep)
        outputfile.write(mermaid_diagram_name + os.linesep)
        for mermaid_string_output in value["strings"]:
            outputfile.write(mermaid_string_output + os.linesep)
